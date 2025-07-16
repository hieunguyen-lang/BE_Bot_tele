import requests
import re
import random
from lxml import html
from datetime import datetime,timedelta
from sqlalchemy.orm import Session
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.sql import func
from sqlalchemy.dialects.mysql import insert
from sqlalchemy import cast, INT
from ..models.hoa_don_models import HoaDon
from ..models.hoa_don_momo_model import HoaDonDien
from ..models.hoa_don_doiung_model import DoiUng
from ..schemas.hoadon_schemas import HoaDonOut,HoaDonUpdate,HoaDonCreate
from .. schemas.hoadon_dien_schemas import HoaDonDienOut
from .. schemas.doiung_schemas import DoiUngOut
from ..models import User, UserRole, hoa_don_models
from ..auth import verify_password, get_password_hash
from fastapi import HTTPException, status
from fastapi.responses import StreamingResponse
from collections import defaultdict
from typing import List, Dict
from sqlalchemy import asc ,desc
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from ..helpers import helper
async def get_hoa_don_stats(db, current_user=User):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not have permission to perform this action."
        )

    # ✅ Gộp: count(*) + count(distinct batch_id) + sum tong_so_tien + sum tien_phi
    stmt = select(
        func.count().label("total_records"),
        func.count(func.distinct(hoa_don_models.HoaDon.batch_id)).label("total_batches"),
        func.sum(hoa_don_models.HoaDon.tong_so_tien).label("total_amount"),
        func.sum(hoa_don_models.HoaDon.phi_per_bill).label("total_fee")
    )

    result = await db.execute(stmt)
    row = result.fetchone()

    return {
        "totalRecords": row.total_records,
        "totalBatches": row.total_batches,
        "totalAmount": int(row.total_amount or 0),
        "totalFee": int(row.total_fee or 0)
    }

async def get_hoa_don_stats_hoa_don(so_hoa_don,so_lo,tid,mid,nguoi_gui,ten_khach,so_dien_thoai,ngay_giao_dich,db, current_user):
    
    from_ = datetime.strptime(ngay_giao_dich, "%Y-%m-%d") if ngay_giao_dich else None
    to_plus_1 = from_ + timedelta(days=1) if from_ else None
    # Tạo base query
    query = select(
        func.count().label("total_records"),
        func.count(func.distinct(hoa_don_models.HoaDon.batch_id)).label("total_batches"),
        func.sum(hoa_don_models.HoaDon.tong_so_tien).label("total_amount"),
        func.sum(hoa_don_models.HoaDon.phi_per_bill).label("total_fee")
    )

    if current_user.role != UserRole.ADMIN:
        query = query.where(hoa_don_models.HoaDon.nguoi_gui == current_user.username)
    # Áp dụng filter nếu có
    if so_hoa_don:
        query = query.where(hoa_don_models.HoaDon.so_hoa_don.contains(so_hoa_don))
    if so_lo:
        query = query.where(hoa_don_models.HoaDon.so_lo.contains(so_lo))
    if tid:
        query = query.where(hoa_don_models.HoaDon.tid.contains(tid))
    if mid:
        query = query.where(hoa_don_models.HoaDon.mid.contains(mid))
    if nguoi_gui:
        query = query.where(hoa_don_models.HoaDon.nguoi_gui.contains(nguoi_gui))
    if ten_khach:
        query = query.where(hoa_don_models.HoaDon.ten_khach.contains(ten_khach))
    if so_dien_thoai:
        query = query.where(hoa_don_models.HoaDon.so_dien_thoai.contains(so_dien_thoai))
    if ngay_giao_dich:
        query = query.where(hoa_don_models.HoaDon.created_at >= from_,
                            hoa_don_models.HoaDon.created_at <=  to_plus_1)

    result = await db.execute(query)
    row = result.fetchone()

    
    return {
        "totalRecords": row.total_records,
        "totalBatches": row.total_batches,
        "totalAmount": int(row.total_amount or 0),
        "totalFee": int(row.total_fee or 0)
    }    


async def get_hoa_don_grouped(page, page_size, db, filters=None,current_user=User):
    
    base_query = select(hoa_don_models.HoaDon)
    # 2. Nếu không phải admin → chỉ được xem hóa đơn của mình
    if current_user.role != UserRole.ADMIN:
        base_query = base_query.where(hoa_don_models.HoaDon.nguoi_gui == current_user.username)
    # Áp dụng filters
    if filters:
        if filters.get("so_hoa_don"):
            base_query = base_query.where(hoa_don_models.HoaDon.so_hoa_don.contains(filters["so_hoa_don"]))
        if filters.get("so_lo"):
            base_query = base_query.where(hoa_don_models.HoaDon.so_lo.contains(filters["so_lo"]))
        if filters.get("tid"):
            base_query = base_query.where(hoa_don_models.HoaDon.tid.contains(filters["tid"]))
        if filters.get("mid"):
            base_query = base_query.where(hoa_don_models.HoaDon.mid.contains(filters["mid"]))
        if filters.get("nguoi_gui"):
            base_query = base_query.where(hoa_don_models.HoaDon.nguoi_gui.contains(filters["nguoi_gui"]))
        if filters.get("ten_khach"):
            base_query = base_query.where(hoa_don_models.HoaDon.ten_khach.contains(filters["ten_khach"]))
        if filters.get("so_dien_thoai"):
            base_query = base_query.where(hoa_don_models.HoaDon.so_dien_thoai.contains(filters["so_dien_thoai"]))
        if filters.get("ngay_giao_dich"):
            from_ = datetime.strptime(filters.get("ngay_giao_dich"), "%Y-%m-%d") if filters.get("ngay_giao_dich") else None
            to_plus_1 = from_ + timedelta(days=1) if from_ else None
            base_query = base_query.where(hoa_don_models.HoaDon.created_at >= from_,
                                        hoa_don_models.HoaDon.created_at <=  to_plus_1)

    # 1. Lấy danh sách batch_id (phân trang) với filter
    sub = (
        select(
            hoa_don_models.HoaDon.batch_id,
            func.min(hoa_don_models.HoaDon.thoi_gian).label("min_time")
        )
        .where(*base_query._where_criteria)  # sử dụng filter có sẵn
        .group_by(hoa_don_models.HoaDon.batch_id)
        .order_by(desc("min_time"))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .subquery()
    )

    # 2. Lấy batch_id từ subquery
    stmt_batch_ids = select(sub.c.batch_id)
    result = await db.execute(stmt_batch_ids)
    batch_ids = [row[0] for row in result.fetchall()]

    # 3. Tổng số batch_id (không cần offset/limit)
    stmt_total = (
        select(func.count())
        .select_from(
            select(hoa_don_models.HoaDon.batch_id)
            .where(*base_query._where_criteria)
            .distinct()
            .subquery()
        )
    )
    total_result = await db.execute(stmt_total)
    total = total_result.scalar()

    # 4. Lấy record theo batch_id
    stmt_records = base_query.where(hoa_don_models.HoaDon.batch_id.in_(batch_ids))
    result = await db.execute(stmt_records)
    records = result.scalars().all()

    # 5. Nhóm lại
    grouped = defaultdict(list)
    for r in records:
        masked_so_the = None
        if r.so_the and len(r.so_the) >= 4:
            masked_so_the = "*" * (len(r.so_the) - 4) + r.so_the[-4:]

        hoa_don_dict = r.__dict__.copy()
        hoa_don_dict["so_the"] = masked_so_the
        grouped[r.batch_id].append(HoaDonOut(**hoa_don_dict))

    data = [
        {"batch_id": batch_id, "records": grouped[batch_id]}
        for batch_id in batch_ids
    ]

    return {"total": total, "data": data}


async def get_hoa_don_dien_grouped(page, page_size, db, filters=None,current_user=User):
   
    base_query = select(HoaDonDien)
    # 2. Nếu không phải admin → chỉ được xem hóa đơn của mình
    if current_user.role != UserRole.ADMIN:
        base_query = base_query.where(HoaDonDien.nguoi_gui == current_user.username)
    # Áp dụng filters
    if filters:
        if filters.get("ma_giao_dich"):
            base_query = base_query.where(HoaDonDien.ma_giao_dich.contains(filters["ma_giao_dich"]))
        if filters.get("ma_khach_hang"):
            base_query = base_query.where(HoaDonDien.ma_khach_hang.contains(filters["ma_khach_hang"]))
        if filters.get("ten_zalo"):
            base_query = base_query.where(HoaDonDien.ten_zalo.contains(filters["ten_zalo"]))
        if filters.get("nguoi_gui"):
            base_query = base_query.where(HoaDonDien.nguoi_gui.contains(filters["nguoi_gui"]))
        # Thêm filter theo thời gian
        if filters.get("from_date"):
            from_date = datetime.strptime(
                filters["from_date"], "%Y-%m-%d"
            ).replace(hour=0, minute=0, second=0, microsecond=0)
            base_query = base_query.where(HoaDonDien.update_at >= from_date)

        if filters.get("to_date"):
            to_date = datetime.strptime(
                filters["to_date"], "%Y-%m-%d"
            ).replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)
            base_query = base_query.where(HoaDonDien.update_at < to_date)


    # 1. Lấy danh sách batch_id (phân trang) với filter
    sub = (
        select(
            HoaDonDien.batch_id,
            func.max(HoaDonDien.update_at).label("max_time")
        )
        .where(*base_query._where_criteria)
        .group_by(HoaDonDien.batch_id)
        .order_by(desc("max_time"))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .subquery()
    )

    # 2. Lấy batch_id từ subquery
    stmt_batch_ids = select(sub.c.batch_id)
    result = await db.execute(stmt_batch_ids)
    batch_ids = [row[0] for row in result.fetchall()]

    # 3. Tổng số batch_id (không cần offset/limit)
    stmt_total = (
        select(func.count())
        .select_from(
            select(HoaDonDien.batch_id)
            .where(*base_query._where_criteria)
            .distinct()
            .subquery()
        )
    )
    total_result = await db.execute(stmt_total)
    total = total_result.scalar()

    # 4. Lấy record theo batch_id
    stmt_records = base_query.where(HoaDonDien.batch_id.in_(batch_ids))
    result = await db.execute(stmt_records)
    records = result.scalars().all()

    # 5. Nhóm lại
    grouped = defaultdict(list)
    for r in records:
        masked_so_the = None

        hoa_don_dict = r.__dict__.copy()
        hoa_don_dict["so_the"] = masked_so_the
        grouped[r.batch_id].append(HoaDonDienOut(**hoa_don_dict))

    data = [
        {"batch_id": batch_id, "records": grouped[batch_id]}
        for batch_id in batch_ids
    ]

    return {"total": total, "data": data}

async def get_hoa_don_dien_stats(db, filters=None, current_user=User):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not have permission to perform this action."
        )

    base_query = select(
        func.count().label("total_records"),
        func.sum(HoaDonDien.so_tien).label("total_amount"),  # ✅ bỏ cast
        func.sum(HoaDonDien.phi_cong_ty_thu).label("total_fee")
    )

    if filters:
        if filters.get("ma_giao_dich"):
            base_query = base_query.where(HoaDonDien.ma_giao_dich.contains(filters["ma_giao_dich"]))
        if filters.get("ma_khach_hang"):
            base_query = base_query.where(HoaDonDien.ma_khach_hang.contains(filters["ma_khach_hang"]))
        if filters.get("ten_zalo"):
            base_query = base_query.where(HoaDonDien.ten_zalo.contains(filters["ten_zalo"]))
        if filters.get("nguoi_gui"):
            base_query = base_query.where(HoaDonDien.nguoi_gui.contains(filters["nguoi_gui"]))
        if filters.get("from_date"):
            from_date = datetime.strptime(
                filters["from_date"], "%Y-%m-%d"
            ).replace(hour=0, minute=0, second=0, microsecond=0)
            base_query = base_query.where(HoaDonDien.update_at >= from_date)

        if filters.get("to_date"):
            to_date = datetime.strptime(
                filters["to_date"], "%Y-%m-%d"
            ).replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)
            base_query = base_query.where(HoaDonDien.update_at < to_date)

    result = await db.execute(base_query)
    row = result.first()

    return {
        "total": row.total_records or 0,
        "totalAmount": float(row.total_amount) if row.total_amount else 0,
        "total_fee": float(row.total_fee)  if row.total_fee else 0,
    }


async def create_hoa_don_dien(db, hoa_don,redis):
    # Kiểm tra trùng mã giao dịch
    result = await db.execute(
        select(HoaDonDien).where(HoaDonDien.ma_giao_dich == hoa_don.ma_giao_dich)
    )
    if result.scalar():
        raise HTTPException(status_code=400, detail="Mã giao dịch đã tồn tại")
    
    hoa_don_data = hoa_don.dict()
    # Nếu không có batch_id hoặc batch_id rỗng thì tự động tạo
    if not hoa_don_data.get("batch_id"):
        batch_id = "_".join([
            str(hoa_don_data.get("ten_khach_hang", "")).strip(),
            str(hoa_don_data.get("ma_khach_hang", "")).strip(),
            str(hoa_don_data.get("dia_chi", "")).strip(),
            str(hoa_don_data.get("so_tien", "")).strip(),
            str(hoa_don_data.get("ma_giao_dich", "")).strip()
        ])
        hoa_don_data["batch_id"] = batch_id
    key_join={
        "ten_khach_hang": hoa_don_data.ten_khach_hang,
        "ma_khach_hang": hoa_don_data.ma_khach_hang,
        "dia_chi": hoa_don_data.dia_chi,
        "so_tien": hoa_don_data.so_tien,
        "ma_giao_dich": hoa_don_data.ma_giao_dich,
    }
    hoa_don_data.key_redis= helper.generate_invoice_dien(key_join)
    obj = HoaDonDien(**hoa_don_data)
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    await redis.sadd("momo_invoices", hoa_don.key_redis)
    return obj

async def update_hoa_don_dien(db, hoa_don,id,redis):
    result = await db.execute(select(HoaDonDien).where(HoaDonDien.id == id))
    hoa_don_data = result.scalar_one_or_none()
    if not hoa_don_data:
        raise HTTPException(status_code=404, detail="Không tìm thấy hóa đơn")
    key_join={
        "ten_khach_hang": hoa_don_data.ten_khach_hang,
        "ma_khach_hang": hoa_don_data.ma_khach_hang,
        "dia_chi": hoa_don_data.dia_chi,
        "so_tien": hoa_don_data.so_tien,
        "ma_giao_dich": hoa_don_data.ma_giao_dich,
    }
    hoa_don_data.key_redis= helper.generate_invoice_dien(key_join)
    for k, v in hoa_don.dict(exclude_unset=True).items():
        setattr(hoa_don_data, k, v)
    await db.commit()
    await db.refresh(hoa_don_data)
    await redis.sadd("momo_invoices", hoa_don.key_redis)
    return hoa_don_data

async def delete_hoa_don_dien(db, id,redis):
    result = await db.execute(select(HoaDonDien).where(HoaDonDien.id == id))
    obj = result.scalar_one_or_none()
    if not obj:
        raise HTTPException(status_code=404, detail="Không tìm thấy hóa đơn")
    await db.delete(obj)
    await db.commit()
    if obj.key_redis:
        await redis.srem("momo_invoices", obj.key_redis)
    return

async def delete_hoa_don_dien_batch(db, batch_id,redis):
    result = await db.execute(select(HoaDonDien).where(HoaDonDien.batch_id == batch_id))
    db_hoa_don_list = result.scalars().all()
    if not db_hoa_don_list:
        raise HTTPException(status_code=404, detail="Batch không có hóa đơn nào!")
    # Đếm số lượng đã xóa
    count_deleted = 0
    # Xóa từng hóa đơn
    for hoa_don in db_hoa_don_list:
        await db.delete(hoa_don)
        count_deleted += 1

        # Nếu có key Redis, xóa key
        if hoa_don.key_redis:
            await redis.srem("momo_invoices", hoa_don.key_redis)
    
    # Commit transaction
    await db.commit()

    return {
        "ok": True,
        "deleted": count_deleted,
        "batch_id": batch_id
    }
    





async def get_doi_ung_flat(page, page_size, db, filters=None, current_user=User):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not have permission to perform this action."
        )
    # Tạo base query với filters
    base_query = select(DoiUng)

    if filters:
        if filters.get("ma_giao_dich"):
            base_query = base_query.where(DoiUng.ma_giao_dich.contains(filters["ma_giao_dich"]))
        if filters.get("ma_khach_hang"):
            base_query = base_query.where(DoiUng.ma_khach_hang.contains(filters["ma_khach_hang"]))
        if filters.get("ten_khach_hang"):
            base_query = base_query.where(DoiUng.ten_khach_hang.contains(filters["ten_khach_hang"]))
        if filters.get("ten_zalo"):
            base_query = base_query.where(DoiUng.ten_zalo == filters["ten_zalo"])
        if filters.get("nguoi_gui"):
            base_query = base_query.where(DoiUng.nguoi_gui.contains(filters["nguoi_gui"]))
        if filters.get("from_date"):
            from_date = datetime.strptime(filters["from_date"], "%Y-%m-%d").date()
            base_query = base_query.where(DoiUng.thoi_gian >= from_date)
        if filters.get("to_date"):
            to_date = datetime.strptime(filters["to_date"], "%Y-%m-%d").date() + timedelta(days=1)
            base_query = base_query.where(DoiUng.thoi_gian < to_date)

    # Tổng số bản ghi thỏa mãn filter
    stmt_total = select(func.count()).select_from(base_query.subquery())
    total_result = await db.execute(stmt_total)
    total = total_result.scalar()

    # Thêm phân trang & sắp xếp
    stmt_records = (
        base_query
        .order_by(desc(DoiUng.thoi_gian))
        .offset((page - 1) * page_size)
        .limit(page_size)
    )

    result = await db.execute(stmt_records)
    records = result.scalars().all()

    # Format ra list
    data = []
    for r in records:
        hoa_don_dict = r.__dict__.copy()
        hoa_don_dict["so_the"] = None  # mask hoặc xử lý nếu cần
        data.append(DoiUngOut(**hoa_don_dict))

    return {
        "total": total,
        "data": data
    }

async def get_doi_ung_stats(db, filters=None, current_user=User):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not have permission to perform this action."
        )

    base_query = select(
        func.count().label("total_records"),
        func.sum(DoiUng.so_tien.cast(INT)).label("total_amount")
    )

    if filters:
        if filters.get("ma_giao_dich"):
            base_query = base_query.where(DoiUng.ma_giao_dich.contains(filters["ma_giao_dich"]))
        if filters.get("ma_khach_hang"):
            base_query = base_query.where(DoiUng.ma_khach_hang.contains(filters["ma_khach_hang"]))
        if filters.get("ten_khach_hang"):
            base_query = base_query.where(DoiUng.ten_khach_hang.contains(filters["ten_khach_hang"]))
        if filters.get("ten_zalo"):
            base_query = base_query.where(DoiUng.ten_zalo == filters["ten_zalo"])
        if filters.get("nguoi_gui"):
            base_query = base_query.where(DoiUng.nguoi_gui.contains(filters["nguoi_gui"]))
        if filters.get("from_date"):
            from_date = datetime.strptime(filters["from_date"], "%Y-%m-%d").date()
            base_query = base_query.where(DoiUng.thoi_gian >= from_date)
        if filters.get("to_date"):
            to_date = datetime.strptime(filters["to_date"], "%Y-%m-%d").date() + timedelta(days=1)
            base_query = base_query.where(DoiUng.thoi_gian < to_date)

    result = await db.execute(base_query)
    row = result.first()

    return {
        "total": row.total_records or 0,
        "totalAmount": int(row.total_amount) if row.total_amount else 0
    }


async def create_hoa_don(db, hoa_don, current_user,redis):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not have permission to perform this action."
        )
    
    # Validation các field
    validation_errors = []
    
    # 1. Validate ngày giao dịch
    if hoa_don.ngay_giao_dich:
        try:
            datetime.strptime(hoa_don.ngay_giao_dich, '%Y-%m-%d')
        except ValueError:
            validation_errors.append("Ngày giao dịch không đúng định dạng YYYY-MM-DD")
    
    # 2. Validate giờ giao dịch
    if hoa_don.gio_giao_dich:
        if not re.match(r'^([01]?[0-9]|2[0-3]):[0-5][0-9]$', hoa_don.gio_giao_dich):
            validation_errors.append("Giờ giao dịch không đúng định dạng HH:MM")
    
    # 3. Validate số tiền
    if hoa_don.tong_so_tien:
        try:
            amount = int(hoa_don.tong_so_tien)
            if amount <= 0:
                validation_errors.append("Tổng số tiền phải lớn hơn 0")
        except ValueError:
            validation_errors.append("Tổng số tiền phải là số nguyên")
    
    # 4. Validate phí
    if hoa_don.tien_phi:
        try:
            fee = int(hoa_don.tien_phi)
            if fee < 0:
                validation_errors.append("Phí không được âm")
        except ValueError:
            validation_errors.append("Phí phải là số nguyên")
    
    # 5. Validate CK vào/ra
    if hoa_don.ck_vao:
        try:
            ck_vao = int(hoa_don.ck_vao)
            if ck_vao < 0:
                validation_errors.append("CK vào không được âm")
        except ValueError:
            validation_errors.append("CK vào phải là số nguyên")
    
    if hoa_don.ck_ra:
        try:
            ck_ra = int(hoa_don.ck_ra)
            if ck_ra < 0:
                validation_errors.append("CK ra không được âm")
        except ValueError:
            validation_errors.append("CK ra phải là số nguyên")
    
    # 6. Validate số điện thoại
    if hoa_don.so_dien_thoai:
        if not re.match(r'^[0-9]{10,11}$', hoa_don.so_dien_thoai):
            validation_errors.append("Số điện thoại không hợp lệ (10-11 số)")
    
    # 7. Validate số thẻ (nếu có)
    if hoa_don.so_the:
        if not re.match(r'^[0-9]{4,19}$', hoa_don.so_the):
            validation_errors.append("Số thẻ không hợp lệ")
    
    # 8. Validate TID/MID
    if hoa_don.tid and len(hoa_don.tid) > 50:
        validation_errors.append("TID quá dài (tối đa 50 ký tự)")
    
    if hoa_don.mid and len(hoa_don.mid) > 50:
        validation_errors.append("MID quá dài (tối đa 50 ký tự)")
    
    # 9. Validate tên khách
    if hoa_don.ten_khach and len(hoa_don.ten_khach.strip()) == 0:
        validation_errors.append("Tên khách không được để trống")
    
    # 10. Validate người gửi
    if hoa_don.nguoi_gui and len(hoa_don.nguoi_gui.strip()) == 0:
        validation_errors.append("Người gửi không được để trống")
    
    # Nếu có lỗi validation, trả về tất cả lỗi
    if validation_errors:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"message": "Validation failed", "errors": validation_errors}
        )
    key_create ={
        'sdt': hoa_don.so_dien_thoai,
        'so_hoa_don': hoa_don.so_hoa_don,
        'gio_giao_dich': hoa_don.gio_giao_dich,
        'so_lo': hoa_don.so_lo,
        'tong_so_tien': hoa_don.tong_so_tien

    }
    hoa_don.key_redis =helper.generate_invoice_key_simple(key_create,hoa_don.ngan_hang)
    
    # Tạo hóa đơn mới
    try:
        db_hoa_don = HoaDon(**hoa_don.dict())
        db.add(db_hoa_don)
        await db.commit()
        await db.refresh(db_hoa_don)
        await redis.sadd("processed_invoices", hoa_don.key_redis)
        return db_hoa_don
    except IntegrityError as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Lỗi khi tạo hóa đơn (có thể trùng lặp dữ liệu)"
        )
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Lỗi server khi tạo hóa đơn"
        )

async def update_hoa_don(
    hoa_don_id, 
    hoa_don,
    db,
    current_user,
    redis
):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not have permission to perform this action."
        )
    # Lấy hóa đơn
    stmt = select(HoaDon).where(HoaDon.id == hoa_don_id)
    result = await db.execute(stmt)
    db_hoa_don = result.scalar_one_or_none()
    
    if not db_hoa_don:
        raise HTTPException(status_code=404, detail="Hóa đơn không tồn tại")
    key_create ={
        'sdt': hoa_don.so_dien_thoai,
        'so_hoa_don': hoa_don.so_hoa_don,
        'gio_giao_dich': hoa_don.gio_giao_dich,
        'so_lo': hoa_don.so_lo,
        'tong_so_tien': hoa_don.tong_so_tien

    }
    # Cập nhật fields
    update_data = hoa_don.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_hoa_don, field, value)
    
    await db.commit()
    await db.refresh(db_hoa_don)
    
    hoa_don.key_redis =helper.generate_invoice_key_simple(key_create,hoa_don.ngan_hang)
    await redis.sadd("processed_invoices", hoa_don.key_redis)
    return db_hoa_don

async def delete_hoa_don(
    hoa_don_id: int, 
    db,
    current_user,
    redis,
):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not have permission to perform this action."
        )
    # Lấy hóa đơn
    stmt = select(HoaDon).where(HoaDon.id == hoa_don_id)
    result = await db.execute(stmt)
    db_hoa_don = result.scalar_one_or_none()
    
    if not db_hoa_don:
        raise HTTPException(status_code=404, detail="Hóa đơn không tồn tại")
    
    # Xóa hóa đơn
    await db.delete(db_hoa_don)
    await db.commit()
    #Xóa key redis
    if db_hoa_don.key_redis:
        await redis.srem("processed_invoices", db_hoa_don.key_redis)
    return {"ok": True}

async def delete_hoa_don_batch_id(
    batch_id: str, 
    db,
    current_user,
    redis,
):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not have permission to perform this action."
        )
    
    # Lấy danh sách hóa đơn theo batch_id
    stmt = select(HoaDon).where(HoaDon.batch_id == batch_id)
    result = await db.execute(stmt)
    db_hoa_don_list = result.scalars().all()
    
    if not db_hoa_don_list:
        raise HTTPException(status_code=404, detail="Batch không có hóa đơn nào!")

    # Đếm số lượng đã xóa
    count_deleted = 0

    # Xóa từng hóa đơn
    for hoa_don in db_hoa_don_list:
        await db.delete(hoa_don)
        count_deleted += 1

        # Nếu có key Redis, xóa key
        if hoa_don.key_redis:
            await redis.srem("processed_invoices", hoa_don.key_redis)
    
    # Commit transaction
    await db.commit()

    return {
        "ok": True,
        "deleted": count_deleted,
        "batch_id": batch_id
    }

async def export_hoa_don_excel(
   page, page_size, db, filters=None,current_user=User     
):
    data = await get_hoa_don_grouped(page, page_size, db, filters,current_user)
    all_rows = []
    for group in data["data"]:
        #batch_id = group["batch_id"]
        records = group["records"]
        # 👉 Merge KẾT TOÁN từ tổng các tong_so_tien
        tong_cong = sum(int(r.tong_so_tien or 0) for r in records)
        for r in records:
            row = {
                "ngay_giao_dich": r.ngay_giao_dich,
                "nguoi_gui": r.nguoi_gui,
                "ten_khach": r.ten_khach,
                "so_dien_thoai": r.so_dien_thoai,
                "type_dao_rut": r.type_dao_rut,
                "ket_toan": tong_cong,
                "so_the": r.so_the,
                "tid": r.tid,
                "so_lo": r.so_lo,
                "so_hoa_don": r.so_hoa_don,
                "gio_giao_dich": r.gio_giao_dich,
                "ten_may_pos": r.ten_may_pos,
                "tong_so_tien": r.tong_so_tien,
                "phan_tram_phi": r.phan_tram_phi,
                "tien_phi": r.tien_phi,
                "ck_ra": r.ck_ra,
                "ck_vao": r.ck_vao,
                "stk_khach": r.stk_khach,
                "stk_cty": r.stk_cty,
                "dia_chi": r.dia_chi,
                "caption_goc": r.caption_goc,
                "ly_do": r.ly_do,
                "tinh_trang": r.tinh_trang
            }

            all_rows.append(row)

    # ✅ Tạo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Hóa đơn"

    headers = [
        "STT",
        "NGÀY",
        "NGƯỜI GỬI",
        "TÊN KHÁCH",
        "SĐT KHÁCH",
        "ĐÁO / RÚT",
        "KẾT TOÁN",
        "SỐ THẺ",
        "TID",
        "SỐ LÔ",
        "SỐ HÓA ĐƠN",
        "GIỜ GIAO DỊCH",
        "POS",
        "SỐ TIỀN",
        "Phí %",
        "PHÍ DV",
        "CK ra",
        "CK vào",
        "STK KHÁCH",
        "STK CÔNG TY",
        "ĐỊA CHỈ",
        "NOTE GỐC",
        "LÝ DO",
        "TÌNH TRẠNG",
    ]

    ws.append(headers)

    for idx, r in enumerate(all_rows, 1):
        ws.append([
            idx,
            r.get("ngay"),
            r.get("nguoi_gui"),
            r.get("ten_khach"),
            r.get("sdt_khach"),
            r.get("loai"),
            r.get("ket_toan"),
            r.get("so_the"),
            r.get("tid"),
            r.get("so_lo"),
            r.get("so_hoa_don"),
            r.get("gio"),
            r.get("ten_pos"),
            r.get("so_tien"),
            r.get("phan_tram_phi"),
            r.get("phi_dv"),
            r.get("ck_ra"),
            r.get("ck_vao"),
            r.get("stk_khach"),
            r.get("stk_cty"),
            r.get("dia_chi"),
            r.get("caption_goc"),
            r.get("ly_do"),
            r.get("tinh_trang"),
        ])


    # Co giãn cột
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=hoa_don.xlsx"}
    )
